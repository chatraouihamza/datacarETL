from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.proxy import Proxy, ProxyType
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
import time
import pandas as pd
import openpyxl

#config openpyxl
notebook=openpyxl.load_workbook("data_details.xlsx")
sheetnames=notebook.active

columns_name = {"Marque" : '', "Couleur" : '', "Modele" : '', "Kilometrage" : '', "Boite de vitesse" : '',
                    "Annee" : '', "Prix TTC" : '', "Puissance Fiscale" : '', "Type de vehicule" : '',"Certificat CRIT'AIR":'', "Co2" : '',
                    "Puissance Reelle" : '', "Type" : '', "Carburant" : '',
                    "Portes" : '', "Places" : '', "Ville" : '', "Garentie" : '',"Mise en circulation":'','Référence':''}

last_row=sheetnames.max_row+1

sheetnames.append(list(columns_name.keys()))
notebook.save("data_details.xlsx")
# Configuration du webdriver
#browser = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))

chrome_option=webdriver.ChromeOptions()
browser=webdriver.Chrome(options=chrome_option)

df=pd.read_csv("link_price.csv")
#browser.get("https://www.autosphere.fr/recherche?chaine=dacia&market=VO&page=10&critaire_checked[]=year&critaire_checked[]=discount&critaire_checked[]=emission_co2")
size=df.shape #rows is 1944 # and df is 1944
for i in range(23,size[0]):   
  url=df.iloc[i,0]
  try:
    browser.get(url);
    browser.delete_all_cookies();
#main=WebDriverWait(browser,10).until(EC.presence_of_element_located((By.CLASS_NAME,'accordion visible-tablet visible-phone')))
    main=WebDriverWait(browser,3).until(EC.presence_of_element_located((By.XPATH,'.//*[@id="menu_mobile"]')))
    parent=main.find_element(By.XPATH,'//*[@id="caract"]/div[1]/div[2]')  #row-fluid description_vehicule
#nom_element = parent[0].get_attribute('class')
#Afficher le nom de l'élément
#print("Nom de l'élément :", nom_element)

    span=parent.find_elements(By.CLASS_NAME,'valeur')
    datas=[];
    for data in span: #les 20 donnees
       texte=data.text
       datas.append(texte)  
#for r in range(df.shape[0]):

    for j,valeur in enumerate(datas,start=1) : 
      sheetnames.cell(row=last_row,column=j,value=valeur)
      notebook.save("data_details.xlsx")
    last_row+=1  
  except TimeoutException:
    # Si l'élément n'est pas trouvé dans le délai imparti, passez au lien suivant
    print("L'élément n'est pas trouvé sur la page. Passage au lien suivant...")
    continue
   
browser.quit()
    






